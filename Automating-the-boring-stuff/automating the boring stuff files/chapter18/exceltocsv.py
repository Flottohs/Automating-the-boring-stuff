
import openpyxl, csv, os

# Loop through all Excel files in the current working directory.

for excel_file in os.listdir('.'):
    if excel_file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excel_file)
        value = 0
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            
            csv_name = f"{os.path.splitext(excel_file)[0]}_{value}.csv"

            
            csvfile = open(csv_name, 'w', newline='')
            csvwrite= csv.writer(csvfile)
            for row_num in range(1, sheet.max_row + 1):
                row_data = []
                for col_num in range(1,sheet.max_column + 1):
                    row_data.append(sheet.cell(row=row_num, column=col_num).value)
                csvwrite.writerow(row_data)
            value +=1
            csvfile.close()
        