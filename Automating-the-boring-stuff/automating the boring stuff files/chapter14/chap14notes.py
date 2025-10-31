# ---------------------------
# READING EXCEL FILES
# ---------------------------

# Use openpyxl to read .xlsx files
import openpyxl

# Load workbook
wb = openpyxl.load_workbook('example.xlsx')  # Load Excel file as a workbook object
type(wb)

# Get sheet names
print(wb.sheetnames)

# Select a sheet
sheet = wb['sheet1']
print(sheet.title)

# Active sheet
another_sheet = wb.active
print(another_sheet.title)

# Accessing cells
cell = sheet['A1']
print(cell.value)
print(f'row {cell.row}, column {cell.column} is {cell.value}')
print(f'cell {cell.coordinate} is {cell.value}')
print(sheet['C1'].value)

# Using row/column indices
sheet.cell(row=1, column=2).value

# Iterating over rows
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# Maximum rows and columns
print(sheet.max_row)
print(sheet.max_column)

# Column letters and indices
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))  # A
print(get_column_letter(27)) # AA
print(column_index_from_string('AA'))  # 27

# Access a range of cells
sheet['A1':'C3']
for row in sheet['A1':'C3']:
    for cell in row:
        print(cell.coordinate, cell.value)
    print('---END OF ROW---')

# Access an entire column (convert to list first)
for cell_obj in list(sheet.columns)[1]:
    print(cell_obj.value)

# ---------------------------
# CREATING & SAVING EXCEL FILES
# ---------------------------

wb = openpyxl.Workbook()  # Creates a new workbook
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
wb.sheetnames  # Check sheet names

# Creating new sheets
wb.create_sheet()  # Add at end
wb.create_sheet(index=0, title='first sheet')
wb.create_sheet(index=2, title='second sheet')

# Deleting sheets
del wb['second sheet']
wb.sheetnames

# Writing to cells
sheet = wb['first sheet']
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)

# ---------------------------
# STYLING CELLS
# ---------------------------

from openpyxl.styles import Font, colors

# Font styles
italic_font = Font(size=24, italic=True)
bold_font = Font(name='Times New Roman', bold=True)

sheet['A1'].font = italic_font
sheet['A1'] = 'Hello, world!'
wb.save('styles3.xlsx')

# ---------------------------
# FORMULAS
# ---------------------------

sheet['B9'] = '=SUM(B1:B8)'  # Set formula
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula3.xlsx')

# To get only values, not formulas
wb = openpyxl.load_workbook('writeFormula3.xlsx', data_only=True)

# ---------------------------
# ADJUSTING ROWS AND COLUMNS
# ---------------------------

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions3.xlsx')

# ---------------------------
# MERGING & UNMERGING CELLS
# ---------------------------

sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.unmerge_cells('A1:D3')

# ---------------------------
# FREEZING PANES
# ---------------------------

# Freeze everything above/to the left of the cell
sheet.freeze_panes = 'A1'  # Example: freeze first row & column
sheet.freeze_panes = None  # Unfreeze

# ---------------------------
# CREATING CHARTS
# ---------------------------

# Add some data
for i in range(1, 11):
    sheet['A' + str(i)] = i * i

# Create chart reference
ref_obj = openpyxl.chart.Reference(sheet, 1, 1, 1, 10)
series_obj = openpyxl.chart.Series(ref_obj, title='First series')
chart_obj = openpyxl.chart.BarChart()
chart_obj.title = 'My Chart'
chart_obj.append(series_obj)
sheet.add_chart(chart_obj, 'C5')




#questions-------
# 1.  What does the openpyxl.load_workbook() function return?

#  2.  What does the wb.sheetnames workbook attribute contain?

#  3.  How would you retrieve the Worksheet object for a sheet named 'Sheet1'?

#  4.  How would you retrieve the Worksheet object for the workbook’s active sheet?

#  5.  How would you retrieve the value in cell C5?

#  6.  How would you set the value in cell C5 to "Hello"?

#  7.  How would you retrieve the cell’s row and column as integers?

#  8.  What do the sheet.max_column and sheet.max_row sheet attributes hold, and what is the data type of these attributes?

#  9.  If you needed to get the integer index for column 'M', what function would you need to call?

#10.  If you needed to get the string name for row 14, what function would you need to call?

#11.  How can you retrieve a tuple of all the Cell objects from A1 to F1?

#12.  How would you save the workbook to the filename example3.xlsx?

#13.  How do you set a formula in a cell?

#14.  If you want to retrieve the result of a cell’s formula instead of the cell’s formula itself, what must you do first?

#15.  How would you set the height of row 5 to 100?

#16.  How would you hide column C?

#17.  What is a freeze pane?

#18.  What five functions and methods do you have to call to create a bar chart?


#answers-------
#1.  It returns a Workbook object representing the Excel file.
#2.  It contains a list of the names of all the sheets in the workbook.
#3.  You would use wb['Sheet1'] to retrieve the Worksheet object for
#    the sheet named 'Sheet1'.
#4.  You would use wb.active to retrieve the Worksheet object for
#    the workbook’s active sheet.
#5.  You would use sheet['C5'].value to retrieve the value in
#    cell C5.
#6.  You would use sheet['C5'] = "Hello" to set the value in
#    cell C5 to "Hello".
#7.  You would use cell.row and cell.column to retrieve the cell’s
#    row and column as integers.
#8.  They hold the maximum number of columns and rows in the sheet,
#    respectively, and their data type is int.
#9.  You would call the column_index_from_string() function.
#10.  You would call the get_column_letter() function.
#11.  You would use sheet['A1':'F1'] to retrieve a tuple of all
#      the Cell objects from A1 to F1.
#12.  You would use wb.save('example3.xlsx') to save the workbook
#      to the filename example3.xlsx.
#13.  You would set a formula in a cell by assigning a string
#      starting with '=' to the cell, e.g., sheet['A1'] = '=SUM(B1:B5)'.
#14.  You must load the workbook with the data_only=True argument,
#      e.g., openpyxl.load_workbook('filename.xlsx', data_only=True).
#15.  You would use sheet.row_dimensions[5].height = 100 to set
#      the height of row 5 to 100.
#16.  You would use sheet.column_dimensions['C'].hidden = True to hide column C.
#17.  A freeze pane is a feature that allows you to keep certain rows
#      or columns visible while scrolling through the rest of the sheet.
#18.  You have to call the following functions and methods to create
#      a bar chart:
#      - openpyxl.chart.Reference() to create a reference to the data.
#      - openpyxl.chart.Series() to create a series from the reference.
#      - openpyxl.chart.BarChart() to create a BarChart object.
#      - chart_obj.append() to add the series to the chart.
#      - sheet.add_chart() to add the chart to the worksheet at a specified location.   
# ---------------------------



